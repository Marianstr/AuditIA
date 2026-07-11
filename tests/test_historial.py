import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import json
import io
import openpyxl


def _seed_historial(entradas):
    with open("historial.json", "w", encoding="utf-8") as f:
        json.dump(entradas, f)


AUDITORIA_BASE = {
    "fecha": "01/01/2026 10:00",
    "tipo": "floristerias",
    "ciudad": "madrid",
    "zona": "",
    "total": 1,
    "calientes": 1,
    "tibios": 0,
    "frios": 0,
    "usuario": "admin",
    "resultados": [{
        "nombre": "Panadería López", "direccion": "Calle Falsa 123", "telefono": "666111222",
        "web": "https://panaderialopez.com", "rating": 4.5, "cantidad_resenas": 10,
        "score": 80, "clasificacion": "caliente", "servicios": ["Diseño y desarrollo web"]
    }]
}


# ---------- GET /historial ----------

def test_historial_vacio_devuelve_lista_vacia(logged_in_client):
    resp = logged_in_client.get("/historial")
    assert resp.status_code == 200
    assert resp.get_json() == []


def test_historial_lista_propias_auditorias_invertidas(logged_in_client):
    primera = {**AUDITORIA_BASE, "fecha": "01/01/2026 10:00"}
    segunda = {**AUDITORIA_BASE, "fecha": "02/01/2026 10:00"}
    _seed_historial([primera, segunda])

    resp = logged_in_client.get("/historial")
    datos = resp.get_json()
    assert len(datos) == 2
    assert datos[0]["fecha"] == "02/01/2026 10:00"
    assert datos[0]["indice"] == 1
    assert datos[1]["indice"] == 0
    assert datos[0]["proyecto"] == ""


def test_historial_filtra_por_usuario(logged_in_client):
    mia = {**AUDITORIA_BASE, "usuario": "admin"}
    ajena = {**AUDITORIA_BASE, "usuario": "otro"}
    _seed_historial([mia, ajena])

    resp = logged_in_client.get("/historial")
    datos = resp.get_json()
    assert len(datos) == 1
    assert datos[0]["indice"] == 0


# ---------- GET /historial/resultados/<indice> ----------

def test_historial_resultados_devuelve_resultados_de_la_auditoria(logged_in_client):
    _seed_historial([AUDITORIA_BASE])
    resp = logged_in_client.get("/historial/resultados/0")
    assert resp.status_code == 200
    assert resp.get_json() == AUDITORIA_BASE["resultados"]


def test_historial_resultados_indice_inexistente_404(logged_in_client):
    _seed_historial([])
    resp = logged_in_client.get("/historial/resultados/5")
    assert resp.status_code == 404
    assert resp.get_json() == {"error": "Auditoría no encontrada"}


def test_historial_resultados_de_otro_usuario_404(logged_in_client):
    _seed_historial([{**AUDITORIA_BASE, "usuario": "otro"}])
    resp = logged_in_client.get("/historial/resultados/0")
    assert resp.status_code == 404


def test_historial_resultados_sin_archivo_404(logged_in_client):
    resp = logged_in_client.get("/historial/resultados/0")
    assert resp.status_code == 404


# ---------- POST /historial/borrar ----------

def test_borrar_historial_uno_propio(logged_in_client):
    primera = {**AUDITORIA_BASE, "fecha": "primera"}
    segunda = {**AUDITORIA_BASE, "fecha": "segunda"}
    _seed_historial([primera, segunda])

    resp = logged_in_client.post("/historial/borrar", json={"indice": 0})
    assert resp.status_code == 200

    with open("historial.json", encoding="utf-8") as f:
        restante = json.load(f)
    assert len(restante) == 1
    assert restante[0]["fecha"] == "segunda"


def test_borrar_historial_no_afecta_auditoria_ajena(logged_in_client):
    _seed_historial([{**AUDITORIA_BASE, "usuario": "otro"}])
    resp = logged_in_client.post("/historial/borrar", json={"indice": 0})
    assert resp.status_code == 200

    with open("historial.json", encoding="utf-8") as f:
        restante = json.load(f)
    assert len(restante) == 1


def test_borrar_historial_todo_solo_afecta_al_usuario(logged_in_client):
    mia = {**AUDITORIA_BASE, "usuario": "admin"}
    ajena = {**AUDITORIA_BASE, "usuario": "otro"}
    _seed_historial([mia, ajena])

    resp = logged_in_client.post("/historial/borrar", json={"todo": True})
    assert resp.status_code == 200

    with open("historial.json", encoding="utf-8") as f:
        restante = json.load(f)
    assert len(restante) == 1
    assert restante[0]["usuario"] == "otro"


def test_borrar_historial_sin_archivo_404(logged_in_client):
    resp = logged_in_client.post("/historial/borrar", json={"indice": 0})
    assert resp.status_code == 404
    assert resp.get_json() == {"ok": False}


# ---------- GET /descargar-excel/<indice> ----------

def test_descargar_excel_devuelve_archivo_para_auditoria_propia(logged_in_client):
    _seed_historial([AUDITORIA_BASE])
    resp = logged_in_client.get("/descargar-excel/0")
    assert resp.status_code == 200
    assert resp.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "leads_floristerias_madrid.xlsx" in resp.headers.get("Content-Disposition", "")

    wb = openpyxl.load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["Nombre", "Direccion", "Telefono", "Web", "Rating", "Resenas", "Score", "Clasificacion", "Servicios recomendados"]
    assert ws[2][0].value == "Panadería López"


def test_descargar_excel_indice_inexistente_404(logged_in_client):
    _seed_historial([])
    resp = logged_in_client.get("/descargar-excel/5")
    assert resp.status_code == 404
    assert "no encontrada" in resp.get_data(as_text=True)


def test_descargar_excel_de_otro_usuario_404(logged_in_client):
    _seed_historial([{**AUDITORIA_BASE, "usuario": "otro"}])
    resp = logged_in_client.get("/descargar-excel/0")
    assert resp.status_code == 404


def test_descargar_excel_sin_archivo_404(logged_in_client):
    resp = logged_in_client.get("/descargar-excel/0")
    assert resp.status_code == 404
